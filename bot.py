# =========================================================
# TELEGRAM BOT ДЛЯ СНТ
# =========================================================
# ФУНКЦИИ:
# - Авторизация по номеру телефона
# - Сохранение номера участка
# - Проверка долгов через Google Sheets
# - Красивый раздел взносов + QR
# - Открытие ворот через звонок
# - Чат СНТ
# - Предложения председателю
# - Голосования
# - Управление голосованиями
# - Модерация чата
# - Система предупреждений
# =========================================================

import os
import re
import uuid
import logging
import aiohttp
import aiosqlite
import tempfile
import openpyxl

from datetime import datetime, timedelta
from dotenv import load_dotenv

from telegram import (
    Update,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ChatPermissions,
    InputFile
)

from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters
)

# =========================================================
# ENV
# =========================================================

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

PAYMENT_LINK = os.getenv("PAYMENT_LINK")

CHAT_LINK = os.getenv("CHAT_LINK")

GOOGLE_SHEETS_ID = os.getenv(
    "GOOGLE_SHEETS_ID"
)

GATE_PHONE = os.getenv("GATE_PHONE")

SNT_CHAT_ID = int(
    os.getenv("SNT_CHAT_ID", "0")
)

ADMIN_ID = int(
    os.getenv("ADMIN_ID", "0")
)

# =========================================================
# LOGGING
# =========================================================

os.makedirs("logs", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(
            "logs/bot.log",
            encoding="utf-8"
        ),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

# =========================================================
# СОСТОЯНИЯ
# =========================================================

(
    ASK_PHONE,
    ASK_PLOT,
    MAIN_MENU,

    ASK_SUGGESTION,

    POLL_MENU,
    CREATE_POLL_QUESTION,
    CREATE_POLL_OPTIONS

) = range(7)

# =========================================================
# RATE LIMIT
# =========================================================

LAST_GATE_OPEN = {}

# =========================================================
# ГОЛОСОВАНИЯ
# =========================================================

ACTIVE_POLLS = {}

USER_POLLS = {}

# =========================================================
# МАТ
# =========================================================

BAD_WORDS = [
    "хуй",
    "пизд",
    "еб",
    "бля",
    "сука",
    "нахуй",
    "мудак",
    "долбоеб",
    "гандон",
    "уеб"
]

# =========================================================
# НОРМАЛИЗАЦИЯ ТЕЛЕФОНА
# =========================================================

def normalize_phone(phone: str) -> str:

    phone = (
        phone.replace(" ", "")
        .replace("-", "")
        .replace("(", "")
        .replace(")", "")
    )

    if phone.startswith("8"):
        phone = "+7" + phone[1:]

    elif phone.startswith("7"):
        phone = "+" + phone

    return phone

# =========================================================
# ПРОВЕРКА МАТА
# =========================================================

def contains_bad_words(text: str) -> bool:

    if not text:
        return False

    text = text.lower()

    for word in BAD_WORDS:

        pattern = rf"\b{re.escape(word)}\w*\b"

        if re.search(pattern, text):
            return True

    return False

# =========================================================
# RATE LIMIT ВОРОТ
# =========================================================

def can_open_gate(user_id: int):

    now = datetime.now()

    if user_id in LAST_GATE_OPEN:

        delta = now - LAST_GATE_OPEN[user_id]

        if delta < timedelta(seconds=30):
            return False

    LAST_GATE_OPEN[user_id] = now

    return True

# =========================================================
# БАЗА ДАННЫХ
# =========================================================

async def init_db():

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        await db.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phone TEXT UNIQUE,
                name TEXT,
                telegram_id INTEGER,
                plot TEXT
            )
        """)

        await db.execute("""
            CREATE TABLE IF NOT EXISTS warnings (
                user_id INTEGER PRIMARY KEY,
                warnings INTEGER DEFAULT 0
            )
        """)

        await db.commit()

# =========================================================
# USERS
# =========================================================

async def get_user_by_phone(phone: str):

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        cursor = await db.execute(
            """
            SELECT id, phone, name
            FROM users
            WHERE phone = ?
            """,
            (phone,)
        )

        return await cursor.fetchone()

async def update_telegram_id(
    phone: str,
    telegram_id: int
):

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        await db.execute(
            """
            UPDATE users
            SET telegram_id = ?
            WHERE phone = ?
            """,
            (telegram_id, phone)
        )

        await db.commit()

async def update_user_plot(
    phone: str,
    plot: str
):

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        await db.execute(
            """
            UPDATE users
            SET plot = ?
            WHERE phone = ?
            """,
            (plot, phone)
        )

        await db.commit()

async def get_user_by_telegram_id(
    telegram_id: int
):

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        cursor = await db.execute(
            """
            SELECT phone, name, plot
            FROM users
            WHERE telegram_id = ?
            """,
            (telegram_id,)
        )

        return await cursor.fetchone()

# =========================================================
# WARNINGS
# =========================================================

async def get_user_warnings(user_id: int):

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        cursor = await db.execute(
            """
            SELECT warnings
            FROM warnings
            WHERE user_id = ?
            """,
            (user_id,)
        )

        row = await cursor.fetchone()

        if row:
            return row[0]

        return 0

async def add_warning(user_id: int):

    warnings = await get_user_warnings(
        user_id
    )

    warnings += 1

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        await db.execute("""
            INSERT INTO warnings(
                user_id,
                warnings
            )
            VALUES(?, ?)

            ON CONFLICT(user_id)
            DO UPDATE SET
            warnings=excluded.warnings
        """, (user_id, warnings))

        await db.commit()

    return warnings

# =========================================================
# DOWNLOAD EXCEL
# =========================================================

async def download_excel():

    try:

        url = (
            f"https://docs.google.com/spreadsheets/d/"
            f"{GOOGLE_SHEETS_ID}/export?format=xlsx"
        )

        async with aiohttp.ClientSession() as session:

            async with session.get(url) as response:

                if response.status != 200:
                    return None

                content = await response.read()

        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as tmp:

            tmp.write(content)

            return tmp.name

    except Exception as e:

        logger.error(
            f"download_excel: {e}"
        )

        return None

# =========================================================
# МОЙ ДОЛГ
# =========================================================

async def get_personal_debt(plot: str):

    try:

        path = await download_excel()

        if not path:
            return None

        wb = openpyxl.load_workbook(path)

        ws = wb.active

        rows = list(
            ws.iter_rows(values_only=True)
        )

        headers = [
            str(h).strip().lower()
            for h in rows[0]
        ]

        idx_plot = headers.index(
            "номер участка"
        )

        idx_required = headers.index(
            "сумма взноса"
        )

        idx_paid = headers.index(
            "фактическая сумма"
        )

        for row in rows[1:]:

            row_plot = row[idx_plot]

            if isinstance(row_plot, float):
                row_plot = str(int(row_plot))
            else:
                row_plot = str(row_plot).strip()

            if row_plot == str(plot):

                required = float(
                    row[idx_required] or 0
                )

                paid = float(
                    row[idx_paid] or 0
                )

                debt = int(required - paid)

                wb.close()

                os.remove(path)

                return max(0, debt)

        wb.close()

        os.remove(path)

        return 0

    except Exception as e:

        logger.error(
            f"personal debt: {e}"
        )

        return None

# =========================================================
# ДОЛЖНИКИ
# =========================================================

async def load_debtors_from_excel():

    try:

        path = await download_excel()

        if not path:
            return "❌ Ошибка загрузки файла"

        wb = openpyxl.load_workbook(path)

        ws = wb.active

        rows = list(
            ws.iter_rows(values_only=True)
        )

        headers = [
            str(h).strip().lower()
            for h in rows[0]
        ]

        idx_plot = headers.index(
            "номер участка"
        )

        idx_required = headers.index(
            "сумма взноса"
        )

        idx_paid = headers.index(
            "фактическая сумма"
        )

        debtors = []

        for row in rows[1:]:

            try:

                plot_value = row[idx_plot]

                if isinstance(plot_value, float):
                    plot = str(int(plot_value))
                else:
                    plot = str(plot_value).strip()

                required = float(
                    row[idx_required] or 0
                )

                paid = float(
                    row[idx_paid] or 0
                )

                debt = int(required - paid)

                if debt > 0:

                    debtors.append(
                        (plot, debt)
                    )

            except:
                continue

        wb.close()

        os.remove(path)

        if not debtors:
            return "✅ Должников нет"

        debtors.sort(
            key=lambda x: x[1],
            reverse=True
        )

        msg = (
            "⚠️ СПИСОК ДОЛЖНИКОВ СНТ\n\n"
        )

        for plot, debt in debtors:

            msg += (
                f"🏠 Участок: {plot}\n"
                f"💰 Долг: {debt:,} ₽\n\n"
            )

        return msg

    except Exception as e:

        logger.error(
            f"debtors: {e}"
        )

        return "❌ Ошибка обработки файла"

# =========================================================
# МОДЕРАЦИЯ
# =========================================================

async def moderate_chat(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    if not update.message:
        return

    if update.message.chat_id != SNT_CHAT_ID:
        return

    if not update.message.text:
        return

    text = update.message.text

    if not contains_bad_words(text):
        return

    user = update.effective_user

    user_id = user.id

    try:
        await update.message.delete()
    except:
        pass

    warnings = await add_warning(user_id)

    username = (
        f"@{user.username}"
        if user.username
        else user.first_name
    )

    if warnings == 1:

        await context.bot.send_message(
            chat_id=SNT_CHAT_ID,
            text=(
                f"⚠️ {username}, предупреждение за мат."
            )
        )

    elif warnings >= 2:

        until_date = (
            datetime.now() +
            timedelta(hours=3)
        )

        await context.bot.restrict_chat_member(
            chat_id=SNT_CHAT_ID,
            user_id=user_id,
            permissions=ChatPermissions(
                can_send_messages=False
            ),
            until_date=until_date
        )

# =========================================================
# START
# =========================================================

async def start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    contact_button = KeyboardButton(
        text="📱 Поделиться номером",
        request_contact=True
    )

    keyboard = [[contact_button]]

    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        resize_keyboard=True
    )

    await update.message.reply_text(
        (
            "🏡 ДОБРО ПОЖАЛОВАТЬ В СНТ\n\n"
            "Для продолжения подтвердите номер телефона."
        ),
        reply_markup=reply_markup
    )

    return ASK_PHONE

# =========================================================
# PHONE
# =========================================================

async def ask_phone(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    contact = update.message.contact

    if not contact:

        await update.message.reply_text(
            "❌ Используйте кнопку ниже."
        )

        return ASK_PHONE

    phone = normalize_phone(
        contact.phone_number
    )

    user = await get_user_by_phone(phone)

    if not user:

        await update.message.reply_text(
            (
                "❌ Ваш номер отсутствует "
                "в базе СНТ."
            ),
            reply_markup=ReplyKeyboardRemove()
        )

        return ConversationHandler.END

    await update_telegram_id(
        phone,
        update.effective_user.id
    )

    context.user_data["phone"] = phone

    await update.message.reply_text(
        f"✅ Добро пожаловать, {user[2]}!"
    )

    async with aiosqlite.connect(
        "database.db"
    ) as db:

        cursor = await db.execute(
            """
            SELECT plot
            FROM users
            WHERE phone = ?
            """,
            (phone,)
        )

        row = await cursor.fetchone()

    if row and row[0]:

        return await show_main_menu(
            update,
            context
        )

    await update.message.reply_text(
        "🏠 Введите номер участка."
    )

    return ASK_PLOT

# =========================================================
# УЧАСТОК
# =========================================================

async def ask_plot(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    plot = update.message.text.strip()

    phone = context.user_data.get(
        "phone"
    )

    await update_user_plot(
        phone,
        plot
    )

    await update.message.reply_text(
        f"✅ Участок №{plot} сохранён."
    )

    return await show_main_menu(
        update,
        context
    )

# =========================================================
# ГЛАВНОЕ МЕНЮ
# =========================================================

async def show_main_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    keyboard = [
        ["💰 Взносы", "💳 Мой долг"],
        ["⚠️ Должники", "💬 Чат СНТ"],
        ["🚪 Открыть ворота", "📝 Предложения"],
        ["📊 Голосования"]
    ]

    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        resize_keyboard=True
    )

    context.user_data["current_state"] = MAIN_MENU
    await update.message.reply_text(
        (
            "📋 ГЛАВНОЕ МЕНЮ\n\n"
            "Выберите нужный раздел."
        ),
        reply_markup=reply_markup
    )

    return MAIN_MENU

# =========================================================
# ВЗНОСЫ
# =========================================================

async def show_payment_section(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user = await get_user_by_telegram_id(
        update.effective_user.id
    )

    plot = user[2]

    debt = await get_personal_debt(plot)

    if debt is None:

        await update.message.reply_text(
            "❌ Ошибка загрузки данных."
        )

        return

    if debt <= 0:

        caption = (
            "💚 ВЗНОСЫ СНТ\n\n"
            f"🏠 Участок: {plot}\n\n"
            "✅ Задолженность отсутствует."
        )

    else:

        caption = (
            "💰 ОПЛАТА ВЗНОСОВ СНТ\n\n"
            f"🏠 Участок: {plot}\n\n"
            f"💳 Текущий долг: {debt:,} ₽\n\n"
            "📌 Для оплаты:\n"
            "• Отсканируйте QR-код\n"
            "• Или нажмите кнопку ниже"
        )

    keyboard = [
        [
            InlineKeyboardButton(
                "💳 ОПЛАТИТЬ ВЗНОСЫ",
                url=PAYMENT_LINK
            )
        ]
    ]

    reply_markup = InlineKeyboardMarkup(
        keyboard
    )

    qr_path = "data/payment_qr.png"

    if os.path.exists(qr_path):

        with open(qr_path, "rb") as qr:

            await update.message.reply_photo(
                photo=InputFile(qr),
                caption=caption,
                reply_markup=reply_markup
            )

    else:

        await update.message.reply_text(
            caption,
            reply_markup=reply_markup
        )

# =========================================================
# ПРЕДЛОЖЕНИЯ
# =========================================================

async def start_suggestion(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    await update.message.reply_text(
        (
            "📝 Напишите предложение "
            "председателю СНТ."
        )
    )

    return ASK_SUGGESTION

async def save_suggestion(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    text = update.message.text

    user = update.effective_user

    sender = (
        f"@{user.username}"
        if user.username
        else user.full_name
    )

    message = (
        "📝 НОВОЕ ПРЕДЛОЖЕНИЕ\n\n"
        f"👤 От: {sender}\n"
        f"🆔 ID: {user.id}\n\n"
        f"{text}"
    )

    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=message
    )

    await update.message.reply_text(
        "✅ Предложение отправлено."
    )

    return await show_main_menu(
        update,
        context
    )

# =========================================================
# МЕНЮ ГОЛОСОВАНИЙ
# =========================================================

async def poll_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_id = update.effective_user.id

    keyboard = []

    if user_id in USER_POLLS:

        keyboard.append(
            ["⚙️ Управлять голосованием"]
        )

    else:

        keyboard.append(
            ["➕ Создать голосование"]
        )

    keyboard.append(
        ["📋 Доступные голосования"]
    )

    keyboard.append(
        ["⬅️ Назад"]
    )

    reply_markup = ReplyKeyboardMarkup(
        keyboard,
        resize_keyboard=True
    )

    context.user_data["current_state"] = POLL_MENU
    await update.message.reply_text(
        (
            "📊 РАЗДЕЛ ГОЛОСОВАНИЙ\n\n"
            "Выберите действие."
        ),
        reply_markup=reply_markup
    )

    context.user_data["current_state"] = POLL_MENU

    return POLL_MENU    

# =========================================================
# СОЗДАНИЕ ГОЛОСОВАНИЯ
# =========================================================

async def start_poll_creation(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    await update.message.reply_text(
        (
            "❓ Напишите вопрос голосования.\n\n"
            "Например:\n"
            "Нужно ли установить больше камер?"
        )
    )

    context.user_data["current_state"] = CREATE_POLL_QUESTION

    return CREATE_POLL_QUESTION

# =========================================================
# ВОПРОС ГОЛОСОВАНИЯ
# =========================================================

async def create_poll_question(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    context.user_data["poll_question"] = (
        update.message.text.strip()
    )

    await update.message.reply_text(
        (
            "✍️ Напишите варианты ответа "
            "через запятую.\n\n"
            "Например:\n"
            "Да, Нет, Воздержался"
        )
    )

    context.user_data["current_state"] = CREATE_POLL_OPTIONS

    return CREATE_POLL_OPTIONS

# =========================================================
# СОЗДАНИЕ ВАРИАНТОВ
# =========================================================

async def create_poll_options(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_id = update.effective_user.id

    question = context.user_data.get(
        "poll_question"
    )

    options = [
        x.strip()
        for x in update.message.text.split(",")
        if x.strip()
    ]

    if len(options) < 2:

        await update.message.reply_text(
            "❌ Минимум 2 варианта."
        )

        return CREATE_POLL_OPTIONS

    poll_id = str(uuid.uuid4())[:8]

    ACTIVE_POLLS[poll_id] = {

        "creator_id": user_id,

        "question": question,

        "options": options,

        "votes": {},

        "active": True
    }

    USER_POLLS[user_id] = poll_id

    await update.message.reply_text(
        (
            "✅ Голосование успешно создано.\n\n"
            "Теперь оно отображается "
            "в разделе доступных голосований."
        )
    )

    return await poll_menu(
        update,
        context
    )

# =========================================================
# СПИСОК ГОЛОСОВАНИЙ
# =========================================================

async def show_polls(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    active = []

    for poll_id, poll in ACTIVE_POLLS.items():

        if poll["active"]:
            active.append((poll_id, poll))

    if not active:

        await update.message.reply_text(
            "📭 Доступных голосований нет."
        )

        return POLL_MENU

    for poll_id, poll in active:

        keyboard = []

        for index, option in enumerate(
            poll["options"]
        ):

            keyboard.append([
                InlineKeyboardButton(
                    text=option,
                    callback_data=(
                        f"vote:{poll_id}:{index}"
                    )
                )
            ])

        reply_markup = InlineKeyboardMarkup(
            keyboard
        )

        await update.message.reply_text(
            (
                "📊 ГОЛОСОВАНИЕ\n\n"
                f"❓ {poll['question']}"
            ),
            reply_markup=reply_markup
        )

    return POLL_MENU

# =========================================================
# CALLBACK ГОЛОСОВАНИЯ
# =========================================================

async def vote_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    data = query.data.split(":")

    poll_id = data[1]

    option_index = int(data[2])

    user = query.from_user

    user_id = user.id

    if poll_id not in ACTIVE_POLLS:

        await query.answer(
            "❌ Голосование не найдено.",
            show_alert=True
        )

        return

    poll = ACTIVE_POLLS[poll_id]

    if not poll["active"]:

        await query.answer(
            "❌ Голосование завершено.",
            show_alert=True
        )

        return

    poll["votes"][user_id] = {

        "option": option_index,

        "name": user.full_name
    }

    await query.answer(
        "✅ Спасибо за ваш голос!",
        show_alert=True
    )

# =========================================================
# УПРАВЛЕНИЕ ГОЛОСОВАНИЕМ
# =========================================================

async def manage_poll(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_id = update.effective_user.id

    if user_id not in USER_POLLS:

        await update.message.reply_text(
            "❌ У вас нет активного голосования."
        )

        return POLL_MENU

    poll_id = USER_POLLS[user_id]

    poll = ACTIVE_POLLS[poll_id]

    report = (
        "📊 ВАШЕ ГОЛОСОВАНИЕ\n\n"
        f"❓ {poll['question']}\n\n"
    )

    if not poll["votes"]:

        report += (
            "📭 Пока никто не голосовал.\n"
        )

    else:

        for vote in poll["votes"].values():

            option = poll["options"][
                vote["option"]
            ]

            report += (
                f"👤 {vote['name']} → "
                f"{option}\n"
            )

    keyboard = [
        [
            InlineKeyboardButton(
                "❌ ЗАВЕРШИТЬ ГОЛОСОВАНИЕ",
                callback_data=f"closepoll:{poll_id}"
            )
        ]
    ]

    reply_markup = InlineKeyboardMarkup(
        keyboard
    )

    await update.message.reply_text(
        report,
        reply_markup=reply_markup
    )

    context.user_data["current_state"] = POLL_MENU

    return POLL_MENU

    
# =========================================================
# ЗАВЕРШЕНИЕ ГОЛОСОВАНИЯ
# =========================================================

async def close_poll_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    query = update.callback_query

    await query.answer()

    poll_id = query.data.split(":")[1]

    if poll_id not in ACTIVE_POLLS:
        return

    poll = ACTIVE_POLLS[poll_id]

    poll["active"] = False

    creator_id = poll["creator_id"]

    if creator_id in USER_POLLS:
        del USER_POLLS[creator_id]

    await query.message.reply_text(
        "✅ Голосование завершено."
    )

# =========================================================
# ВОРОТА
# =========================================================

async def open_gate(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    user_id = update.effective_user.id

    if not can_open_gate(user_id):

        await update.message.reply_text(
            (
                "⏳ Повторный звонок "
                "доступен через 30 секунд."
            )
        )

        return MAIN_MENU

    # Telegram Bot API НЕ поддерживает tel:
    # поэтому просто показываем номер

    text = (
        "🚪 ОТКРЫТИЕ ВОРОТ\n\n"
        "📱 Для открытия ворот:\n\n"
        f"☎️ Позвоните на номер:\n"
        f"{GATE_PHONE}\n\n"
        "❗️Функция звонка работает "
        "только в мобильной версии Telegram.\n\n"
        "Скопируйте номер или нажмите на него."
    )

    await update.message.reply_text(text)

    return MAIN_MENU

# =========================================================
# HANDLE MENU
# =========================================================

async def handle_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    text = update.message.text

    # ВЗНОСЫ

    if text == "💰 Взносы":

        await show_payment_section(
            update,
            context
        )

    # МОЙ ДОЛГ

    elif text == "💳 Мой долг":

        user = await get_user_by_telegram_id(
            update.effective_user.id
        )

        plot = user[2]

        debt = await get_personal_debt(plot)

        if debt <= 0:

            await update.message.reply_text(
                (
                    f"🏠 Участок: {plot}\n\n"
                    "✅ Долгов нет."
                )
            )

        else:

            await update.message.reply_text(
                (
                    f"🏠 Участок: {plot}\n\n"
                    f"💰 Долг: {debt:,} ₽"
                )
            )

    # ДОЛЖНИКИ

    elif text == "⚠️ Должники":

        loading = await update.message.reply_text(
            "📄 Загружаем список..."
        )

        msg = await load_debtors_from_excel()

        await loading.delete()

        await update.message.reply_text(msg)

    # ЧАТ

    elif text == "💬 Чат СНТ":

        keyboard = [
            [
                InlineKeyboardButton(
                    "💬 ОТКРЫТЬ ЧАТ",
                    url=CHAT_LINK
                )
            ]
        ]

        reply_markup = InlineKeyboardMarkup(
            keyboard
        )

        await update.message.reply_text(
            (
                "💬 ЧАТ СНТ\n\n"
                "Нажмите кнопку ниже "
                "для перехода в чат."
            ),
            reply_markup=reply_markup
        )

    # ВОРОТА

    elif text == "🚪 Открыть ворота":

        return await open_gate(
            update,
            context
        )

    # ПРЕДЛОЖЕНИЯ

    elif text == "📝 Предложения":

        return await start_suggestion(
            update,
            context
        )

    # ГОЛОСОВАНИЯ

    elif text == "📊 Голосования":

        return await poll_menu(
            update,
            context
        )

    elif text == "➕ Создать голосование":

        return await start_poll_creation(
            update,
            context
        )

    elif text == "📋 Доступные голосования":

        return await show_polls(
            update,
            context
        )

    elif text == "⚙️ Управлять голосованием":

        return await manage_poll(
            update,
            context
        )

    elif text == "⬅️ Назад":

        return await show_main_menu(
            update,
            context
        )

    current_state = context.user_data.get(
        "current_state",
        MAIN_MENU
    )

    return current_state

# =========================================================
# CANCEL
# =========================================================

async def cancel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    await update.message.reply_text(
        "👋 До свидания.",
        reply_markup=ReplyKeyboardRemove()
    )

    return ConversationHandler.END

# =========================================================
# POST INIT
# =========================================================

async def post_init(app: Application):

    await init_db()

    logger.info("Database initialized")

# =========================================================
# MAIN
# =========================================================

def main():

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .post_init(post_init)
        .build()
    )

    # =====================================================
    # CONVERSATION HANDLER
    # =====================================================

    conv_handler = ConversationHandler(

        entry_points=[
            CommandHandler("start", start)
        ],

        states={

            ASK_PHONE: [

                MessageHandler(
                    filters.ALL,
                    ask_phone
                )
            ],

            ASK_PLOT: [

                MessageHandler(
                    filters.TEXT &
                    ~filters.COMMAND,
                    ask_plot
                )
            ],

            MAIN_MENU: [

                MessageHandler(
                    filters.TEXT &
                    ~filters.COMMAND,
                    handle_menu
                )
            ],

            ASK_SUGGESTION: [

                MessageHandler(
                    filters.TEXT &
                    ~filters.COMMAND,
                    save_suggestion
                )
            ],

            POLL_MENU: [

                MessageHandler(
                    filters.TEXT &
                    ~filters.COMMAND,
                    handle_menu
                )
            ],

            CREATE_POLL_QUESTION: [

                MessageHandler(
                    filters.TEXT &
                    ~filters.COMMAND,
                    create_poll_question
                )
            ],

            CREATE_POLL_OPTIONS: [

                MessageHandler(
                    filters.TEXT &
                    ~filters.COMMAND,
                    create_poll_options
                )
            ]
        },

        fallbacks=[
            CommandHandler(
                "cancel",
                cancel
            )
        ]
    )

    # =====================================================
    # МОДЕРАЦИЯ
    # =====================================================

    app.add_handler(
        MessageHandler(
            filters.TEXT &
            ~filters.COMMAND,
            moderate_chat
        ),
        group=0
    )

    # =====================================================
    # ОСНОВНОЙ HANDLER
    # =====================================================

    app.add_handler(
        conv_handler,
        group=1
    )

    # =====================================================
    # CALLBACK ГОЛОСОВАНИЯ
    # =====================================================

    app.add_handler(
        CallbackQueryHandler(
            vote_callback,
            pattern="^vote:"
        )
    )

    app.add_handler(
        CallbackQueryHandler(
            close_poll_callback,
            pattern="^closepoll:"
        )
    )

    logger.info("Bot started")

    app.run_polling()

# =========================================================
# START BOT
# =========================================================

if __name__ == "__main__":
    main()